import requests
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from pathlib import Path


def fetch_option_chain(symbol):
    session = requests.Session()
    headers = {
        'User-Agent': 'Mozilla/5.0',
        'Accept-Language': 'en-US,en;q=0.9'
    }
    session.get("https://www.nseindia.com/option-chain", headers=headers, timeout=10)
    api_url = f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}"
    resp = session.get(api_url, headers=headers, timeout=10)
    data = resp.json()
    return data

def extract_refresh_timestamp(data):
    # Try common fields or extract from any string
    timestamp_str = data['records'].get('time') or data['records'].get('timestamp')
    if not timestamp_str:
        match = re.search(r'\d{2}-[A-Z][a-z]{2}-\d{4} \d{2}:\d{2}:\d{2}', str(data))
        if match:
            timestamp_str = match.group(0)
    if not timestamp_str:
        raise Exception("Timestamp not found.")
    dt = pd.to_datetime(timestamp_str, format="%d-%b-%Y %H:%M:%S", errors="coerce")
    if pd.isnull(dt):
        raise Exception(f"Timestamp format error: {timestamp_str}")
    return dt.strftime("%d%b%Y%H%M%S").lower()

def format_option_chain(data):
    option_data = data['records']['data']
    underlying = data['records']['underlyingValue']
    rows = []
    for row in option_data:
        strike = row.get('strikePrice', '')
        expiry = row.get('expiryDate', '')
        ce = row.get('CE', {})
        pe = row.get('PE', {})
        rows.append([
            ce.get('impliedVolatility', ''),
            ce.get('lastPrice', ''),
            ce.get('totalTradedVolume', ''),
            ce.get('openInterest', ''),
            ce.get('changeinOpenInterest', ''),
            strike,
            pe.get('changeinOpenInterest', ''),
            pe.get('openInterest', ''),
            pe.get('totalTradedVolume', ''),
            pe.get('lastPrice', ''),
            pe.get('impliedVolatility', ''),
            expiry
        ])
    columns = [
        "CALL_IV", "CALL_LTP", "CALL_VOLUME", "CALL_OI", "CALL_CHG_IN_OI",
        "STRIKE",
        "PUT_CHG_IN_OI", "PUT_OI", "PUT_VOLUME", "PUT_LTP", "PUT_IV",
        "EXPIRY_DATE"
    ]
    df = pd.DataFrame(rows, columns=columns)
    df['EXPIRY_DATE'] = pd.to_datetime(df['EXPIRY_DATE'], dayfirst=True, errors='coerce')
    df = df.sort_values(['EXPIRY_DATE', 'STRIKE'], ascending=[True, True])
    df['EXPIRY_DATE'] = df['EXPIRY_DATE'].dt.strftime('%d-%b-%Y')
    return df, underlying

def style_excel(filename, underlying):
    wb = load_workbook(filename)
    ws = wb.active
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'] = "CALLS"
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
    ws['G1'] = "PUTS"
    ws['F1'] = "STRIKE"
    ws['L1'] = "EXPIRY DATE"
    for col in [1,6,7,12]:
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    call_headers = ["IV", "LTP", "VOLUME", "OI", "CHG IN OI"]
    put_headers = ["CHG IN OI", "OI", "VOLUME", "LTP", "IV"]
    for i, txt in enumerate(call_headers):
        ws.cell(row=2, column=1+i).value = txt
    ws['F2'] = "STRIKE"
    for i, txt in enumerate(put_headers):
        ws.cell(row=2, column=7+i).value = txt
    ws['L2'] = "EXPIRY DATE"
    header_fill = PatternFill('solid', fgColor='7a7a7a')
    black_bold_font = Font(bold=True, color='000000')
    for col in range(1, 13):
        for row in [1,2]:
            c = ws.cell(row=row, column=col)
            c.fill = header_fill
            c.font = black_bold_font
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws['A3']
    itm_fill = PatternFill('solid', fgColor='e1e8f6')
    for i in range(3, ws.max_row + 1):
        strike = ws.cell(row=i, column=6).value
        if isinstance(strike, (int, float)):
            if strike < underlying:
                for c in range(1, 6):
                    ws.cell(row=i, column=c).fill = itm_fill
            if strike > underlying:
                for c in range(7, 12):
                    ws.cell(row=i, column=c).fill = itm_fill
    expiry_col = 12
    call_chg_col = 5
    call_vol_col = 3
    put_chg_col = 7
    put_vol_col = 9
    call_max_color = '5df073'
    call_min_color = 'ff1744'
    put_max_color = call_max_color
    put_min_color = call_min_color
    volume_max_color = '3a75d3'
    expiries = sorted({ws.cell(row=i, column=expiry_col).value for i in range(3, ws.max_row+1)})
    for exp in expiries:
        rows = [i for i in range(3, ws.max_row+1) if ws.cell(row=i, column=expiry_col).value == exp]
        vals = [ws.cell(row=i, column=call_chg_col).value for i in rows if isinstance(ws.cell(row=i, column=call_chg_col).value, (int, float))]
        if vals:
            maxv = max(vals)
            minv = min(vals)
            for i in rows:
                val = ws.cell(row=i, column=call_chg_col).value
                if val == maxv:
                    ws.cell(row=i, column=call_chg_col).fill = PatternFill('solid', fgColor=call_max_color)
                elif val == minv:
                    ws.cell(row=i, column=call_chg_col).fill = PatternFill('solid', fgColor=call_min_color)
        vals = [ws.cell(row=i, column=call_vol_col).value for i in rows if isinstance(ws.cell(row=i, column=call_vol_col).value, (int, float)) and ws.cell(row=i, column=call_vol_col).value != 0]
        if vals:
            maxv = max(vals)
            for i in rows:
                val = ws.cell(row=i, column=call_vol_col).value
                if val == maxv:
                    ws.cell(row=i, column=call_vol_col).fill = PatternFill('solid', fgColor=volume_max_color)
        vals = [ws.cell(row=i, column=put_chg_col).value for i in rows if isinstance(ws.cell(row=i, column=put_chg_col).value, (int, float))]
        if vals:
            maxv = max(vals)
            minv = min(vals)
            for i in rows:
                val = ws.cell(row=i, column=put_chg_col).value
                if val == maxv:
                    ws.cell(row=i, column=put_chg_col).fill = PatternFill('solid', fgColor=put_max_color)
                elif val == minv:
                    ws.cell(row=i, column=put_chg_col).fill = PatternFill('solid', fgColor=put_min_color)
        vals = [ws.cell(row=i, column=put_vol_col).value for i in rows if isinstance(ws.cell(row=i, column=put_vol_col).value, (int, float)) and ws.cell(row=i, column=put_vol_col).value != 0]
        if vals:
            maxv = max(vals)
            for i in rows:
                val = ws.cell(row=i, column=put_vol_col).value
                if val == maxv:
                    ws.cell(row=i, column=put_vol_col).fill = PatternFill('solid', fgColor=volume_max_color)
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

def save_to_excel(df, symbol, underlying, timestamp_tag):
    filename = f"{symbol}_option_chain_{timestamp_tag}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    if not os.path.exists(filepath):
        df.to_excel(filepath, index=False)
        style_excel(filepath, underlying)
        # === Added wrapper to allow Flask to call this script and to save next to this file ===
SCRIPT_DIR = Path(__file__).resolve().parent

def run_option_chain_export(out_dir: Path | None = None):
    """
    Fetch NIFTY & BANKNIFTY, write XLSX files and the underlying text files
    into out_dir (default = folder of this file).

    New behavior:
    - If a file with the same timestamp already exists → skip creating/overwriting.
    - If timestamp is new → create a new XLSX file and keep old ones as-is.
    - Underlying text files are always updated.
    """
    base = Path(out_dir) if out_dir else SCRIPT_DIR
    base.mkdir(parents=True, exist_ok=True)

    written = []
    for symbol in ["NIFTY", "BANKNIFTY"]:
        data = fetch_option_chain(symbol)
        df, underlying = format_option_chain(data)

        try:
            timestamp_tag = extract_refresh_timestamp(data)
        except Exception:
            ts = (pd.Timestamp.utcnow().tz_localize("UTC").tz_convert("Asia/Kolkata"))
            timestamp_tag = ts.strftime("%d%b%Y%H%M%S").lower()

        filepath = base / f"{symbol}_option_chain_{timestamp_tag}.xlsx"

        # Only write if this timestamped file does not already exist
        if not filepath.exists():
            df.to_excel(filepath, index=False)
            try:
                style_excel(str(filepath), underlying)
            except Exception:
                # styling failure is non-fatal
                pass

        # Always update underlying text file (latest spot value)
        (base / f"{symbol}_underlying.txt").write_text(f"{underlying:.2f}", encoding="utf-8")

        written.append(str(filepath))

    return {"written": written}
    
def generate_levels_from_latest(out_dir: Path | None = None):
    """
    Build levels_<timestamp>.xlsx from the latest pair of:
      NIFTY_option_chain_<ts>.xlsx and BANKNIFTY_option_chain_<ts>.xlsx

    Rules:
    - Prefer the most recent COMMON <ts> present for both indices.
    - If no common <ts>, fall back to the most recent available <ts> overall.
    - If levels_<ts>.xlsx already exists, skip creating it.
    - Does NOT modify or call any existing logic.
    """
    base = Path(out_dir) if out_dir else Path(__file__).resolve().parent

    n_files = list(base.glob("NIFTY_option_chain_*.xlsx"))
    b_files = list(base.glob("BANKNIFTY_option_chain_*.xlsx"))
    if not n_files and not b_files:
        return {"ok": False, "reason": "no option-chain files found", "path": None}

    ts_rx = re.compile(r"_option_chain_([0-9]{1,2}[A-Za-z]{3}[0-9]{4}[0-9]{6})\.xlsx$", re.IGNORECASE)
    def ts_of(p: Path):
        m = ts_rx.search(p.name)
        return m.group(1) if m else None

    tn = {ts_of(p): p for p in n_files if ts_of(p)}
    tb = {ts_of(p): p for p in b_files if ts_of(p)}

    def ts_to_dt(ts: str):
        # normalize like 26aug2025153000 -> 26Aug2025153000 for parsing
        return pd.to_datetime(ts.title(), format="%d%b%Y%H%M%S", errors="coerce")

    commons = sorted([t for t in (tn.keys() & tb.keys())], key=lambda x: ts_to_dt(x))
    if commons:
        ts = commons[-1]
        p_n, p_b = tn[ts], tb[ts]
    else:
        all_ts = sorted(list(tn.keys() | tb.keys()), key=lambda x: ts_to_dt(x))
        ts = all_ts[-1]
        p_n, p_b = tn.get(ts), tb.get(ts)

    out_xlsx = base / f"levels_{ts}.xlsx"
    if out_xlsx.exists():
        return {"ok": True, "skipped": True, "reason": "levels already exists", "path": str(out_xlsx)}

    # Read styled OC (has 2 header rows created by style_excel)
    def read_oc(path: Path) -> pd.DataFrame:
        if not path or not path.exists():
            return pd.DataFrame(columns=["EXPIRY DATE","STRIKE","LTP_CALL","LTP_PUT"])
        df_raw = pd.read_excel(path, header=[0,1], engine="openpyxl")

        def col_name(col):
            sec = str(col[0]).strip().upper()
            sub = str(col[1]).strip().upper()
            if sec == "CALLS" and sub not in ("STRIKE",""):
                return f"{sub}_CALL"
            if sec == "PUTS" and sub not in ("STRIKE",""):
                return f"{sub}_PUT"
            if sub == "STRIKE" or sec == "STRIKE":
                return "STRIKE"
            if sub == "EXPIRY DATE" or sec == "EXPIRY DATE":
                return "EXPIRY DATE"
            return sub or sec

        df_raw.columns = [col_name(c) for c in df_raw.columns.values]
        # Keep the minimal set needed for S/R levels
        for k in ("EXPIRY DATE","STRIKE","LTP_CALL","LTP_PUT"):
            if k not in df_raw.columns:
                df_raw[k] = pd.NA
        df = df_raw[["EXPIRY DATE","STRIKE","LTP_CALL","LTP_PUT"]].copy()

        df["STRIKE"]   = pd.to_numeric(df["STRIKE"], errors="coerce")
        df["LTP_CALL"] = pd.to_numeric(df["LTP_CALL"], errors="coerce")
        df["LTP_PUT"]  = pd.to_numeric(df["LTP_PUT"],  errors="coerce")
        if "EXPIRY DATE" in df.columns:
            df["EXPIRY DATE"] = (
                pd.to_datetime(df["EXPIRY DATE"], errors="coerce")
                .dt.strftime("%d-%b-%Y")
            )
        return df

    def add_levels(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        ce   = df["LTP_CALL"].fillna(0)
        pe   = df["LTP_PUT"].fillna(0)
        stk  = df["STRIKE"].fillna(0)
        mn   = ce.where(ce <= pe, pe)     # min(CE, PE)
        mx   = ce.where(ce >= pe, pe)     # max(CE, PE)
        tot  = ce.add(pe, fill_value=0)   # CE+PE

        out = df.copy()
        out["S1"] = (stk - mn).round(2)
        out["S2"] = (stk - mx).round(2)
        out["S3"] = (stk - tot).round(2)
        out["R1"] = (stk + mn).round(2)
        out["R2"] = (stk + mx).round(2)
        out["R3"] = (stk + tot).round(2)
        cols = ["EXPIRY DATE","STRIKE","LTP_CALL","LTP_PUT","S1","S2","S3","R1","R2","R3"]
        return out[[c for c in cols if c in out.columns]]

    df_n = add_levels(read_oc(p_n))
    df_b = add_levels(read_oc(p_b))

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        df_n.to_excel(xw, index=False, sheet_name="NIFTY")
        df_b.to_excel(xw, index=False, sheet_name="BANKNIFTY")

    return {"ok": True, "skipped": False, "path": str(out_xlsx), "timestamp": ts}
# --- END NEW ---

if __name__ == "__main__":
    res = run_option_chain_export()
    print("Written:")
    for p in res.get("written", []):
        print(" -", p)
        
    # NEW: also build levels file after writing option chains
    try:
        lev = generate_levels_from_latest()
        print("Levels:", lev)
    except Exception as e:
        print("Levels generation failed:", e)
